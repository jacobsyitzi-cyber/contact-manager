from __future__ import annotations

import csv
import io
import quopri
import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

try:
    import openpyxl
except Exception:
    openpyxl = None

APP_NAME = "YJ - RCUK CONTACT MANAGER V3"
APP_VERSION = "Streamlit Edition 1.0"
COPYRIGHT_TEXT = "© YITZI JACOBS RCUK 2026"


# ---------------------------- Utilities ----------------------------

def safe_read_text(path: str, encodings=("utf-8", "utf-16", "latin-1")) -> str:
    data = Path(path).read_bytes()
    for enc in encodings:
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("latin-1", errors="ignore")


def safe_decode_text_bytes(data: bytes, encodings=("utf-8", "utf-16", "latin-1")) -> str:
    for enc in encodings:
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("latin-1", errors="ignore")


def decode_contact_bytes(data: bytes) -> str:
    for enc in ("utf-8", "utf-16", "utf-16-le", "latin-1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("latin-1", errors="ignore")


def normalize_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def title_case_name(s: str) -> str:
    s = normalize_spaces(s)
    if not s:
        return s
    parts = re.split(r"(\s+)", s.lower())

    def cap(w: str) -> str:
        if not w or w.isspace():
            return w
        return "-".join([p[:1].upper() + p[1:] if p else p for p in w.split("-")])

    return "".join(cap(p) for p in parts)


def strip_trailing_name_numbers(name: str) -> str:
    n = normalize_spaces(name)
    n = re.sub(r"\s*\(\s*\d+\s*\)\s*$", "", n)
    n = re.sub(r"([_\- ]+)\d+\s*$", "", n)
    n = re.sub(r"(\D)\d+\s*$", r"\1", n)
    return normalize_spaces(n)


def digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")


def to_uk_format(num: str) -> str:
    raw = digits_only(num)
    if not raw:
        return ""
    if raw.startswith("44"):
        raw = raw[2:]
    if raw.startswith("0"):
        return raw
    return "0" + raw


def to_international_44(num: str) -> str:
    raw = digits_only(num)
    if not raw:
        return ""
    if raw.startswith("44"):
        return "+44" + raw[2:]
    if raw.startswith("0"):
        return "+44" + raw[1:]
    return "+44" + raw


def unique_preserve(seq: Iterable[str]) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for x in seq:
        x = x.strip()
        if not x:
            continue
        k = x.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(x)
    return out


# ---------------------------- VCF parsing/writing ----------------------------

VCARD_RE = re.compile(r"BEGIN:VCARD\s*(.*?)\s*END:VCARD", re.IGNORECASE | re.DOTALL)


def split_vcards(text: str) -> List[str]:
    cards = []
    for m in VCARD_RE.finditer(text):
        body = m.group(0).strip()
        if body:
            cards.append(body if body.endswith("\n") else body + "\n")
    return cards


def unfold_vcard_lines(card: str) -> List[str]:
    lines = card.splitlines()
    out: List[str] = []
    for ln in lines:
        if ln.startswith((" ", "\t")) and out:
            out[-1] += ln[1:]
        else:
            out.append(ln)
    return out


def _decode_vcard_value(key: str, val: str) -> str:
    k = key.upper()
    v = val
    if "QUOTED-PRINTABLE" in k:
        try:
            b = quopri.decodestring(v.encode("latin-1", errors="ignore"))
            try:
                return b.decode("utf-8", errors="ignore").strip()
            except Exception:
                return b.decode("latin-1", errors="ignore").strip()
        except Exception:
            return v.strip()
    return v.strip()


def parse_vcard_fields(card: str) -> Dict[str, object]:
    fields: Dict[str, object] = {"fn": "", "n": "", "tels": [], "emails": [], "raw": card}
    fn = ""
    nfield = ""
    tels: List[str] = []
    emails: List[str] = []
    for ln in unfold_vcard_lines(card):
        if ":" not in ln:
            continue
        key, val = ln.split(":", 1)
        k = key.upper()
        v = _decode_vcard_value(k, val)
        if k.startswith("FN"):
            fn = v
        elif k.startswith("N"):
            nfield = v
        elif k.startswith("TEL"):
            tels.append(v)
        elif k.startswith("EMAIL"):
            emails.append(v)
    if not fn and nfield:
        parts = nfield.split(";")
        given = parts[1].strip() if len(parts) > 1 else ""
        family = parts[0].strip() if len(parts) > 0 else ""
        fn = normalize_spaces(f"{given} {family}".strip()) or nfield.replace(";", " ").strip()
    fields["fn"] = fn
    fields["n"] = nfield
    fields["tels"] = unique_preserve(tels)
    fields["emails"] = unique_preserve(emails)
    return fields


def _vcf_escape(value: str) -> str:
    return str(value or "").replace("\\", "\\\\").replace("\n", "\\n").strip()


def build_vcard_extended(
    fn: str,
    tels: List[str],
    emails: List[str],
    *,
    first_name: str = "",
    last_name: str = "",
    middle_name: str = "",
    prefix: str = "",
    suffix: str = "",
    company: str = "",
    job_title: str = "",
    address: str = "",
    city: str = "",
    state: str = "",
    postcode: str = "",
    country: str = "",
    website: str = "",
    notes: str = "",
    tels_extra: Optional[List[str]] = None,
    emails_extra: Optional[List[str]] = None,
) -> str:
    lines = ["BEGIN:VCARD", "VERSION:3.0"]

    final_fn = _vcf_escape(fn)
    if not final_fn:
        final_fn = _vcf_escape(" ".join([prefix, first_name, middle_name, last_name, suffix]).strip())
    if final_fn:
        lines.append(f"FN:{final_fn}")

    if any([last_name, first_name, middle_name, prefix, suffix]):
        n_parts = [
            _vcf_escape(last_name),
            _vcf_escape(first_name),
            _vcf_escape(middle_name),
            _vcf_escape(prefix),
            _vcf_escape(suffix),
        ]
        lines.append("N:" + ";".join(n_parts))

    if company:
        lines.append(f"ORG:{_vcf_escape(company)}")
    if job_title:
        lines.append(f"TITLE:{_vcf_escape(job_title)}")

    all_tels = list(tels or []) + list(tels_extra or [])
    for t in unique_preserve([_vcf_escape(x) for x in all_tels if str(x).strip()]):
        lines.append(f"TEL:{t}")

    all_emails = list(emails or []) + list(emails_extra or [])
    for e in unique_preserve([_vcf_escape(x) for x in all_emails if str(x).strip()]):
        lines.append(f"EMAIL:{e}")

    if any([address, city, state, postcode, country]):
        adr_parts = [
            "", "",
            _vcf_escape(address),
            _vcf_escape(city),
            _vcf_escape(state),
            _vcf_escape(postcode),
            _vcf_escape(country),
        ]
        lines.append("ADR:{}".format(";".join(adr_parts)))

    if website:
        lines.append(f"URL:{_vcf_escape(website)}")
    if notes:
        lines.append(f"NOTE:{_vcf_escape(notes)}")

    lines.append("END:VCARD")
    return "\n".join(lines) + "\n"


def build_vcard(fn: str, tels: List[str], emails: List[str]) -> str:
    return build_vcard_extended(fn, tels, emails)


def apply_normalization_to_card(card: str, phone_mode: str, trim_spaces: bool, unify_case: bool, strip_nums: bool) -> str:
    f = parse_vcard_fields(card)
    fn = str(f.get("fn") or "")
    tels = list(f.get("tels") or [])
    emails = list(f.get("emails") or [])

    tels = [t for t in tels if digits_only(t)]

    if trim_spaces:
        fn = normalize_spaces(fn)
        tels = [normalize_spaces(t) for t in tels]
        emails = [normalize_spaces(e) for e in emails]
    if strip_nums:
        fn = strip_trailing_name_numbers(fn)
    if unify_case:
        fn = title_case_name(fn)
    if phone_mode == "uk":
        tels = [to_uk_format(t) for t in tels]
    elif phone_mode == "international":
        tels = [to_international_44(t) for t in tels]

    return build_vcard(fn, tels, emails)


def vcard_key(card: str) -> Tuple[str, str, str]:
    f = parse_vcard_fields(card)
    fn = (str(f.get("fn") or "")).strip().lower()
    email = (",".join(f.get("emails") or [])).strip().lower()
    phone = (",".join(f.get("tels") or [])).strip().lower()
    return (email, phone, fn)


def best_display_name(card: str) -> str:
    f = parse_vcard_fields(card)
    fn = str(f.get("fn") or "").strip()
    if fn:
        return fn
    tels = f.get("tels") or []
    emails = f.get("emails") or []
    return (tels[0] if tels else (emails[0] if emails else ""))


def smart_merge_cards(cards: List[str]) -> str:
    best_fn = ""
    all_tels: List[str] = []
    all_emails: List[str] = []
    for c in cards:
        f = parse_vcard_fields(c)
        fn = str(f.get("fn") or "").strip()
        if fn and len(fn) > len(best_fn):
            best_fn = fn
        all_tels.extend(list(f.get("tels") or []))
        all_emails.extend(list(f.get("emails") or []))
    return build_vcard(best_fn, unique_preserve(all_tels), unique_preserve(all_emails))


# ---------------------------- Importers ----------------------------

def extract_vcards_from_bytes(data: bytes) -> List[str]:
    texts: List[str] = []
    data_no_null = data.replace(b"\x00", b"")
    for b in (data_no_null, data):
        for enc in ("utf-8", "utf-16-le", "utf-16", "latin-1"):
            try:
                texts.append(b.decode(enc, errors="ignore"))
            except Exception:
                continue
    cards: List[str] = []
    seen: Set[str] = set()
    for t in texts:
        for c in split_vcards(t):
            k = c.strip()
            if not k:
                continue
            if k in seen:
                continue
            seen.add(k)
            cards.append(c)
    return cards


def import_from_vcf_file(path: str) -> List[str]:
    text = safe_read_text(path)
    return split_vcards(text)


def import_from_vcf_bytes(data: bytes) -> List[str]:
    text = safe_decode_text_bytes(data)
    return split_vcards(text)


def import_from_csv(path: str) -> Tuple[List[Dict[str, str]], List[str]]:
    text = safe_read_text(path)
    return import_from_csv_text(text)


def import_from_csv_text(text: str) -> Tuple[List[Dict[str, str]], List[str]]:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    rows = [{k or "": (v or "") for k, v in (row or {}).items()} for row in reader]
    headers = [h for h in (reader.fieldnames or [])]
    return rows, headers


def import_from_xlsx(path: str) -> Tuple[List[Dict[str, str]], List[str]]:
    if openpyxl is None:
        raise RuntimeError("XLSX support requires openpyxl. Install: pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    return workbook_to_rows(wb)


def import_from_xlsx_bytes(data: bytes) -> Tuple[List[Dict[str, str]], List[str]]:
    if openpyxl is None:
        raise RuntimeError("XLSX support requires openpyxl. Install: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    return workbook_to_rows(wb)


def workbook_to_rows(wb):
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows_iter)]
    out_rows: List[Dict[str, str]] = []
    for r in rows_iter:
        d = {}
        for i, h in enumerate(headers):
            d[h] = "" if i >= len(r) else ("" if r[i] is None else str(r[i]))
        out_rows.append(d)
    return out_rows, headers


def import_from_nbf_or_zip(path: str) -> Tuple[List[str], List[Tuple[str, int]]]:
    data = Path(path).read_bytes()
    return import_from_nbf_or_zip_bytes(data, Path(path).name)


def import_from_nbf_or_zip_bytes(data: bytes, filename: str = "upload") -> Tuple[List[str], List[Tuple[str, int]]]:
    per: List[Tuple[str, int]] = []
    cards: List[str] = []
    bio = io.BytesIO(data)

    is_zip = False
    try:
        is_zip = zipfile.is_zipfile(bio)
    except Exception:
        is_zip = False

    if is_zip:
        bio.seek(0)
        with zipfile.ZipFile(bio, "r") as z:
            for name in sorted(z.namelist(), key=str.lower):
                if not name.lower().endswith(".vcf"):
                    continue
                try:
                    content = z.read(name)
                    text = decode_contact_bytes(content)
                    vc = split_vcards(text)
                    if not vc:
                        vc = extract_vcards_from_bytes(content)
                    cards.extend(vc)
                    per.append((name, len(vc)))
                except Exception:
                    continue
    else:
        vc = extract_vcards_from_bytes(data)
        cards.extend(vc)
        per.append((filename, len(vc)))

    return cards, per


def import_from_ib(path: str) -> Tuple[List[str], List[Tuple[str, int]]]:
    return import_from_ib_bytes(Path(path).read_bytes(), Path(path).name)


def import_from_ib_bytes(data: bytes, filename: str = "upload") -> Tuple[List[str], List[Tuple[str, int]]]:
    per: List[Tuple[str, int]] = []
    cards: List[str] = []

    vc = extract_vcards_from_bytes(data)
    if vc:
        cards.extend(vc)
        per.append((filename, len(vc)))
        return cards, per

    def extract_utf16le_strings(b: bytes, minlen: int = 2) -> List[Tuple[str, int]]:
        out: List[Tuple[str, int]] = []
        cur: List[str] = []
        start_pos: Optional[int] = None
        for i in range(0, len(b) - 1, 2):
            codept = int.from_bytes(b[i:i + 2], "little")
            if 32 <= codept <= 126:
                if start_pos is None:
                    start_pos = i
                cur.append(chr(codept))
            else:
                if start_pos is not None and len(cur) >= minlen:
                    out.append(("".join(cur), start_pos))
                cur = []
                start_pos = None
        if start_pos is not None and len(cur) >= minlen:
            out.append(("".join(cur), start_pos))
        return out

    def looks_like_name(s: str) -> bool:
        ss = s.strip()
        if not ss or ss.lower().endswith(".ib"):
            return False
        return any(ch.isalpha() for ch in ss) and len(ss) >= 2

    def bcd_phone_near(b: bytes, window: int = 1200) -> str:
        def decode_bcd(sub: bytes) -> Optional[str]:
            digits: List[str] = []
            for byte in sub:
                for nib in (byte & 0x0F, (byte >> 4) & 0x0F):
                    if nib <= 9:
                        digits.append(str(nib))
                    elif nib == 0xF:
                        continue
                    else:
                        return None
            return "".join(digits)

        lim = min(window, len(b))
        for i in range(0, lim - 6):
            for nbytes in range(6, 9):
                if i + nbytes > lim:
                    continue
                d = decode_bcd(b[i:i + nbytes])
                if not d:
                    continue
                if (len(d) == 11 and d.startswith(("07", "01", "02", "03", "08", "09"))) or (len(d) == 12 and d.startswith("44")):
                    return d
        return ""

    pairs: List[Tuple[str, str]] = []
    seen: Set[str] = set()
    for s, pos in extract_utf16le_strings(data, minlen=2):
        name = s.strip()
        if not looks_like_name(name):
            continue
        k = name.lower()
        if k in seen:
            continue
        seen.add(k)
        phone_raw = bcd_phone_near(data[pos:pos + 1200], window=1200)
        pairs.append((name, phone_raw))

    made = 0
    for name, phone_raw in pairs:
        tels: List[str] = []
        if phone_raw:
            tels = ["+" + phone_raw] if phone_raw.startswith("44") else [phone_raw]
        cards.append(build_vcard(name, tels, []))
        made += 1

    per.append((filename, made))
    return cards, per


# ---------------------------- Spreadsheet mapping ----------------------------

TARGET_FIELDS = [
    "Ignore",
    "Full Name",
    "First Name",
    "Last Name",
    "Middle Name",
    "Name Prefix",
    "Name Suffix",
    "Phone",
    "Phone 2",
    "Phone 3",
    "Email",
    "Email 2",
    "Company",
    "Job Title",
    "Address",
    "City",
    "State/County",
    "Postcode",
    "Country",
    "Website",
    "Notes",
]


def guess_target(header: str) -> str:
    h = (header or "").strip().lower()
    if any(k in h for k in ["full name", "contact name", "display name"]):
        return "Full Name"
    if h == "name":
        return "Full Name"
    if any(k in h for k in ["first name", "forename", "given name"]):
        return "First Name"
    if any(k in h for k in ["last name", "surname", "family name"]):
        return "Last Name"
    if "middle" in h:
        return "Middle Name"
    if "prefix" in h:
        return "Name Prefix"
    if "suffix" in h:
        return "Name Suffix"
    if any(k in h for k in ["mobile", "telephone", "phone", "tel", "number"]):
        if "2" in h or "secondary" in h or "other" in h:
            return "Phone 2"
        if "3" in h or "third" in h:
            return "Phone 3"
        return "Phone"
    if "email" in h or h == "mail":
        if "2" in h or "secondary" in h or "other" in h:
            return "Email 2"
        return "Email"
    if any(k in h for k in ["company", "organisation", "organization", "business"]):
        return "Company"
    if any(k in h for k in ["job", "title", "position", "role"]):
        return "Job Title"
    if h == "address" or "street" in h:
        return "Address"
    if "city" in h or "town" in h:
        return "City"
    if any(k in h for k in ["state", "county", "province", "region"]):
        return "State/County"
    if any(k in h for k in ["postcode", "zip", "postal"]):
        return "Postcode"
    if "country" in h:
        return "Country"
    if any(k in h for k in ["website", "web", "url"]):
        return "Website"
    if any(k in h for k in ["note", "notes", "comment", "comments"]):
        return "Notes"
    return "Ignore"


def build_cards_from_mapped_rows(rows: List[Dict[str, str]], mapping: Dict[str, str]) -> List[str]:
    cards: List[str] = []

    for r in rows:
        field_values: Dict[str, List[str]] = {k: [] for k in TARGET_FIELDS if k != "Ignore"}

        for header, target in mapping.items():
            if target == "Ignore":
                continue
            val = str(r.get(header, "") or "").strip()
            if val:
                field_values.setdefault(target, []).append(val)

        full_name = " ".join(field_values.get("Full Name", [])).strip()
        first_name = " ".join(field_values.get("First Name", [])).strip()
        last_name = " ".join(field_values.get("Last Name", [])).strip()
        middle_name = " ".join(field_values.get("Middle Name", [])).strip()
        prefix = " ".join(field_values.get("Name Prefix", [])).strip()
        suffix = " ".join(field_values.get("Name Suffix", [])).strip()

        if not full_name:
            full_name = " ".join([prefix, first_name, middle_name, last_name, suffix]).strip()

        phones_primary = []
        for key in ("Phone", "Phone 2", "Phone 3"):
            phones_primary.extend(field_values.get(key, []))

        emails_primary = []
        for key in ("Email", "Email 2"):
            emails_primary.extend(field_values.get(key, []))

        card = build_vcard_extended(
            full_name,
            phones_primary,
            emails_primary,
            first_name=first_name,
            last_name=last_name,
            middle_name=middle_name,
            prefix=prefix,
            suffix=suffix,
            company=" ".join(field_values.get("Company", [])).strip(),
            job_title=" ".join(field_values.get("Job Title", [])).strip(),
            address=" ".join(field_values.get("Address", [])).strip(),
            city=" ".join(field_values.get("City", [])).strip(),
            state=" ".join(field_values.get("State/County", [])).strip(),
            postcode=" ".join(field_values.get("Postcode", [])).strip(),
            country=" ".join(field_values.get("Country", [])).strip(),
            website=" ".join(field_values.get("Website", [])).strip(),
            notes=" | ".join(field_values.get("Notes", [])).strip(),
        )
        cards.append(card)

    return cards


# ---------------------------- Deduplication ----------------------------

def group_duplicates(cards: List[str], match_email: bool = True, match_phone: bool = True, match_name: bool = True) -> List[List[int]]:
    by_email: Dict[str, List[int]] = {}
    by_phone: Dict[str, List[int]] = {}
    by_fn: Dict[str, List[int]] = {}

    for i, c in enumerate(cards):
        f = parse_vcard_fields(c)
        emails = [e.lower() for e in (f.get("emails") or []) if e]
        tels = [digits_only(t) for t in (f.get("tels") or []) if t]
        fn = (str(f.get("fn") or "")).strip().lower()

        if match_email and emails:
            by_email.setdefault(emails[0], []).append(i)
        if match_phone and tels:
            by_phone.setdefault(tels[0], []).append(i)
        if match_name and fn:
            by_fn.setdefault(fn, []).append(i)

    groups: List[List[int]] = []
    used: Set[int] = set()

    def add_groups(d: Dict[str, List[int]]):
        nonlocal groups, used
        for _, idxs in d.items():
            if len(idxs) <= 1:
                continue
            g = [x for x in idxs if x not in used]
            if len(g) > 1:
                groups.append(g)
                used.update(g)

    if match_email:
        add_groups(by_email)
    if match_phone:
        add_groups(by_phone)
    if match_name:
        add_groups(by_fn)
    return groups


# ---------------------------- Models and exports ----------------------------

@dataclass
class ImportedData:
    cards: List[str] = field(default_factory=list)
    sources: List[Tuple[str, int]] = field(default_factory=list)
    kind: str = ""
    raw_rows: Optional[List[Dict[str, str]]] = None
    headers: Optional[List[str]] = None
    filename: str = ""


@dataclass
class NormalizationOptions:
    phone_mode: str = "keep"
    trim_spaces: bool = False
    unify_case: bool = False
    strip_name_numbers: bool = False


@dataclass
class ExportBundle:
    files: Dict[str, bytes]
    summary_name: str
    summary_text: str


def apply_normalization_all(cards: List[str], options: NormalizationOptions) -> List[str]:
    return [
        apply_normalization_to_card(
            c,
            options.phone_mode,
            options.trim_spaces,
            options.unify_case,
            options.strip_name_numbers,
        )
        for c in cards
    ]


def export_convert_bundle(imported: ImportedData, options: NormalizationOptions, export_mode: str, batch_size: int = 500) -> ExportBundle:
    normalized = apply_normalization_all(imported.cards, options)
    files: Dict[str, bytes] = {}
    if export_mode == "merged":
        files["contacts.vcf"] = "".join(normalized).encode("utf-8")
    elif export_mode == "per_contact":
        for i, c in enumerate(normalized, start=1):
            files[f"contacts/contact_{i}.vcf"] = c.encode("utf-8")
    else:
        bs = max(1, int(batch_size or 500))
        batch = 1
        for start in range(0, len(normalized), bs):
            part = normalized[start:start + bs]
            files[f"batches/contacts_batch_{batch}.vcf"] = "".join(part).encode("utf-8")
            batch += 1

    summary_text = build_summary(
        exported=len(normalized),
        sources=imported.sources,
        options=options,
        heading=APP_NAME,
    )
    files["RCUK_Contact_Manager_Summary.txt"] = summary_text.encode("utf-8")
    return ExportBundle(files=files, summary_name="RCUK_Contact_Manager_Summary.txt", summary_text=summary_text)


def export_merge_bundle(cards: List[str], options: NormalizationOptions, remove_duplicates: bool, smart_merge: bool) -> ExportBundle:
    all_cards = apply_normalization_all(cards, options)
    if remove_duplicates:
        if smart_merge:
            grouped: Dict[Tuple[str, str, str], List[str]] = {}
            for c in all_cards:
                grouped.setdefault(vcard_key(c), []).append(c)
            final_cards = [smart_merge_cards(v) if len(v) > 1 else v[0] for v in grouped.values()]
        else:
            seen: Set[Tuple[str, str, str]] = set()
            final_cards = []
            for c in all_cards:
                k = vcard_key(c)
                if k in seen:
                    continue
                seen.add(k)
                final_cards.append(c)
    else:
        final_cards = all_cards

    summary_text = build_summary(
        exported=len(final_cards),
        sources=[("Merged VCF input", len(cards))],
        options=options,
        heading=f"{APP_NAME} - Merge",
    )
    return ExportBundle(
        files={
            "merged_contacts.vcf": "".join(final_cards).encode("utf-8"),
            "RCUK_Contact_Manager_Summary.txt": summary_text.encode("utf-8"),
        },
        summary_name="RCUK_Contact_Manager_Summary.txt",
        summary_text=summary_text,
    )


def export_dedupe_bundle(
    cards: List[str],
    options: NormalizationOptions,
    dup_groups: List[List[int]],
    keep_choices: Dict[int, int],
    smart_merge: bool,
) -> ExportBundle:
    if smart_merge:
        removed: Set[int] = set()
        merged: List[str] = []
        for _, grp in enumerate(dup_groups):
            if len(grp) < 2:
                continue
            removed.update(grp)
            merged.append(smart_merge_cards([cards[i] for i in grp]))
        kept = [cards[i] for i in range(len(cards)) if i not in removed]
        out_cards = kept + merged
    else:
        to_remove: Set[int] = set()
        for gi, grp in enumerate(dup_groups):
            keep = keep_choices.get(gi, grp[0])
            for idx in grp:
                if idx != keep:
                    to_remove.add(idx)
        out_cards = [cards[i] for i in range(len(cards)) if i not in to_remove]

    out_cards = apply_normalization_all(out_cards, options)
    summary_text = build_summary(
        exported=len(out_cards),
        sources=[("De-duplicate input", len(cards))],
        options=options,
        heading=f"{APP_NAME} - Deduplicate",
    )
    return ExportBundle(
        files={
            "deduplicated.vcf": "".join(out_cards).encode("utf-8"),
            "RCUK_Contact_Manager_Summary.txt": summary_text.encode("utf-8"),
        },
        summary_name="RCUK_Contact_Manager_Summary.txt",
        summary_text=summary_text,
    )


def build_summary(exported: int, sources: List[Tuple[str, int]], options: NormalizationOptions, heading: str) -> str:
    lines = [
        heading,
        f"Exported: {exported} contacts",
        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "Source(s):",
    ]
    for s, cnt in (sources or []):
        lines.append(f"  - {s}: {cnt}")
    lines.extend([
        "",
        "Normalisation:",
        f"  Phone mode: {options.phone_mode}",
        f"  Trim spaces: {options.trim_spaces}",
        f"  Unify case: {options.unify_case}",
        f"  Remove numbers after names: {options.strip_name_numbers}",
        "",
    ])
    return "\n".join(lines)
